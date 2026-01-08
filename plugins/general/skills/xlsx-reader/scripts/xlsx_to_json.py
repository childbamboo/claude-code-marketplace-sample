#!/usr/bin/env python3
"""
Excel to JSON Converter
Converts Excel (.xlsx) files to JSON format.
"""

import json
import sys
import os
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is not installed.")
    print("Please install it with: pip install openpyxl")
    sys.exit(1)


def excel_to_json(file_path, max_rows=None):
    """
    Convert Excel file to JSON format.

    Args:
        file_path (str): Path to the Excel file
        max_rows (int): Maximum rows to include per sheet (None = all)

    Returns:
        dict: JSON-formatted data
    """
    wb = load_workbook(file_path, data_only=True)
    result = {
        "filename": os.path.basename(file_path),
        "total_sheets": len(wb.sheetnames),
        "sheets": {}
    }

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        max_row = sheet.max_row if sheet.max_row else 0
        max_col = sheet.max_column if sheet.max_column else 0

        if max_row == 0 or max_col == 0:
            result["sheets"][sheet_name] = {
                "dimensions": {"rows": 0, "columns": 0},
                "data": []
            }
            continue

        # Limit rows if specified
        if max_rows and max_row > max_rows:
            actual_rows = max_rows
        else:
            actual_rows = max_row

        # Read header row
        headers = []
        for col_idx in range(1, max_col + 1):
            cell = sheet.cell(row=1, column=col_idx)
            header = str(cell.value) if cell.value else f"column_{col_idx}"
            headers.append(header)

        # Read data rows
        data = []
        for row_idx in range(2, actual_rows + 1):
            row_data = {}
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = cell.value

                # Convert datetime to ISO format string
                if isinstance(value, datetime):
                    value = value.isoformat()
                elif value is None:
                    value = None
                else:
                    # Keep numbers as numbers, convert others to string
                    if not isinstance(value, (int, float)):
                        value = str(value)

                row_data[headers[col_idx - 1]] = value

            data.append(row_data)

        result["sheets"][sheet_name] = {
            "dimensions": {
                "rows": max_row,
                "columns": max_col,
                "rows_included": actual_rows
            },
            "data": data
        }

    wb.close()
    return result


def main():
    """Main entry point for the script."""
    if len(sys.argv) < 2:
        print("Usage: python xlsx_to_json.py <file_path> [max_rows]")
        print("\nExamples:")
        print("  python xlsx_to_json.py data.xlsx")
        print("  python xlsx_to_json.py data.xlsx 100")
        sys.exit(1)

    file_path = sys.argv[1]

    if not os.path.exists(file_path):
        print(f"Error: File not found: {file_path}")
        sys.exit(1)

    # Parse max rows
    max_rows = None
    if len(sys.argv) > 2:
        try:
            max_rows = int(sys.argv[2])
        except ValueError:
            print(f"Warning: Invalid max_rows value: {sys.argv[2]}")

    # Convert to JSON
    print(f"Converting {file_path} to JSON...")
    data = excel_to_json(file_path, max_rows)

    # Output JSON
    json_output = json.dumps(data, ensure_ascii=False, indent=2)
    print(json_output)

    # Summary
    print(f"\n# Conversion Summary", file=sys.stderr)
    print(f"Total sheets: {data['total_sheets']}", file=sys.stderr)
    for sheet_name, sheet_data in data["sheets"].items():
        rows = len(sheet_data["data"])
        print(f"  {sheet_name}: {rows} rows", file=sys.stderr)


if __name__ == "__main__":
    main()
