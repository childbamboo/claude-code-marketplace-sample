#!/usr/bin/env python3
"""
Excel Reader Script
Reads Excel (.xlsx) files and converts to Markdown format.
"""

import sys
import os

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is not installed.")
    print("Please install it with: pip install openpyxl")
    sys.exit(1)


def read_xlsx(file_path, sheet_names=None, max_rows=None):
    """
    Read Excel file and convert to Markdown.

    Args:
        file_path (str): Path to the Excel file
        sheet_names (list): List of sheet names to read (None = all sheets)
        max_rows (int): Maximum rows to read per sheet (None = all rows)

    Returns:
        str: Markdown formatted content
    """
    try:
        # Use normal mode to ensure dimensions are calculated correctly
        workbook = load_workbook(file_path, data_only=True)
        markdown_content = []

        # Add document header
        markdown_content.append(f"# {os.path.basename(file_path)}\n")
        markdown_content.append(f"**Total Sheets:** {len(workbook.sheetnames)}\n")
        markdown_content.append("---\n")

        # Process each sheet
        sheets_to_process = sheet_names if sheet_names else workbook.sheetnames

        for sheet_name in sheets_to_process:
            if sheet_name not in workbook.sheetnames:
                markdown_content.append(f"\n⚠️ Sheet '{sheet_name}' not found\n")
                continue

            sheet = workbook[sheet_name]
            markdown_content.append(f"\n## Sheet: {sheet_name}\n")

            # Get dimensions
            max_row = sheet.max_row if sheet.max_row else 0
            max_col = sheet.max_column if sheet.max_column else 0

            if max_rows and max_row > max_rows:
                markdown_content.append(f"*Showing first {max_rows} of {max_row} rows*\n")
                max_row = max_rows

            markdown_content.append(f"**Dimensions:** {max_row} rows × {max_col} columns\n")

            # Check if sheet is empty
            if max_row == 0 or max_col == 0 or max_row is None or max_col is None:
                markdown_content.append("\n*(Empty sheet)*\n")
                continue

            # Convert to Markdown table
            table_md = convert_sheet_to_markdown(sheet, max_row, max_col)
            markdown_content.append(f"\n{table_md}\n")
            markdown_content.append("\n---\n")

        workbook.close()
        return '\n'.join(markdown_content)

    except FileNotFoundError:
        return f"Error: File not found: {file_path}"
    except Exception as e:
        return f"Error reading Excel file: {str(e)}"


def convert_sheet_to_markdown(sheet, max_row, max_col):
    """
    Convert Excel sheet to Markdown table.

    Args:
        sheet: openpyxl worksheet object
        max_row (int): Maximum row to read
        max_col (int): Maximum column to read

    Returns:
        str: Markdown table
    """
    markdown_lines = []

    # Read all cells
    rows = []
    for row_idx in range(1, max_row + 1):
        row_data = []
        for col_idx in range(1, max_col + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = cell.value

            # Format cell value
            if value is None:
                value = ""
            else:
                value = str(value).strip()

            row_data.append(value)
        rows.append(row_data)

    if not rows:
        return "*(No data)*"

    # Create Markdown table
    # Header row
    header = rows[0]
    markdown_lines.append("| " + " | ".join(header) + " |")

    # Separator row
    markdown_lines.append("| " + " | ".join(["---"] * len(header)) + " |")

    # Data rows
    for row in rows[1:]:
        # Pad row if necessary
        while len(row) < len(header):
            row.append("")
        markdown_lines.append("| " + " | ".join(row[:len(header)]) + " |")

    return '\n'.join(markdown_lines)


def main():
    """Main entry point for the script."""
    if len(sys.argv) < 2:
        print("Usage: python read_xlsx.py <file_path> [sheet_names] [max_rows]")
        print("\nExamples:")
        print("  python read_xlsx.py data.xlsx")
        print("  python read_xlsx.py data.xlsx 'Sheet1,Sheet2'")
        print("  python read_xlsx.py data.xlsx 'Sheet1' 100")
        sys.exit(1)

    file_path = sys.argv[1]

    if not os.path.exists(file_path):
        print(f"Error: File not found: {file_path}")
        sys.exit(1)

    if not file_path.lower().endswith('.xlsx'):
        print("Warning: File does not have .xlsx extension")

    # Parse sheet names
    sheet_names = None
    if len(sys.argv) > 2 and sys.argv[2]:
        sheet_names = [s.strip() for s in sys.argv[2].split(',')]

    # Parse max rows
    max_rows = None
    if len(sys.argv) > 3:
        try:
            max_rows = int(sys.argv[3])
        except ValueError:
            print(f"Warning: Invalid max_rows value: {sys.argv[3]}")

    content = read_xlsx(file_path, sheet_names, max_rows)
    print(content)


if __name__ == "__main__":
    main()
